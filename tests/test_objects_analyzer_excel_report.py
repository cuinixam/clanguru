#!/usr/bin/env python3

from pathlib import Path

from openpyxl import load_workbook

from clanguru.compilation_options_manager import CompileCommand
from clanguru.object_analyzer import ObjectDependencies, ObjectReportData, ObjectsDataExcelReportGenerator, Symbol, SymbolLinkage


def create_object_report_data_for_excel_test(object_name: str, symbols: list[Symbol]) -> ObjectReportData:
    """Helper function to create ObjectReportData for Excel tests."""
    object_path = Path(object_name)
    source_path = object_path.with_suffix(".c")

    # Create mock CompileCommand
    compile_command = CompileCommand(
        directory=Path("/mock/build"),
        file=source_path,
        command=f"gcc -c {source_path} -o {object_path}",
    )

    # Create ObjectDependencies
    object_dependencies = ObjectDependencies(path=object_path, symbols=symbols)

    return ObjectReportData(object_dependencies=object_dependencies, compile_command=compile_command)


def test_excel_report_generator(tmp_path: Path) -> None:
    """Create a sample Excel report with mock object data."""
    # Create sample object data
    obj1 = create_object_report_data_for_excel_test(
        "main.o",
        [
            Symbol("main", SymbolLinkage.LOCAL),
            Symbol("printf", SymbolLinkage.EXTERN),
            Symbol("strlen", SymbolLinkage.EXTERN),
            Symbol("my_function", SymbolLinkage.LOCAL),
        ],
    )

    obj2 = create_object_report_data_for_excel_test(
        "utils.o",
        [
            Symbol("my_function", SymbolLinkage.EXTERN),
            Symbol("utility_func", SymbolLinkage.LOCAL),
            Symbol("malloc", SymbolLinkage.EXTERN),
            Symbol("free", SymbolLinkage.EXTERN),
        ],
    )

    obj3 = create_object_report_data_for_excel_test(
        "math.o",
        [
            Symbol("add", SymbolLinkage.LOCAL),
            Symbol("subtract", SymbolLinkage.LOCAL),
            Symbol("multiply", SymbolLinkage.LOCAL),
            Symbol("sin", SymbolLinkage.EXTERN),
            Symbol("cos", SymbolLinkage.EXTERN),
        ],
    )

    # Create the report generator
    generator = ObjectsDataExcelReportGenerator([obj1, obj2, obj3], create_traceability_matrix=True)

    # Generate the Excel report
    output_file = tmp_path / "objects_report.xlsx"
    generator.generate_report(output_file)

    assert output_file.exists(), f"Report file {output_file} was not created."

    workbook = load_workbook(output_file)
    assert workbook.sheetnames == ["Objects", "Dependency Matrix"]
